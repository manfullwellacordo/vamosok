import pandas as pd
import numpy as np
import os
from pathlib import Path
from datetime import datetime, timedelta
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestClassifier, GradientBoostingRegressor
from sklearn.model_selection import train_test_split
import joblib
import warnings
import matplotlib.pyplot as plt
import seaborn as sns
import jinja2
import base64
from io import BytesIO
import json
import traceback
import xlrd  # Para arquivos .xls
from openpyxl import load_workbook  # Para arquivos .xlsx
import sqlite3
import re

class AnalisadorInteligente:
    def __init__(self):
        self.diretorios = [
            Path('F:/relatoriotest'),
            Path('./data'),
            Path('.')
        ]
        
        self.arquivos = {
            'julio': '(JULIO) LISTAS INDIVIDUAIS.xlsx',
            'leandro': '(LEANDRO_ADRIANO) LISTAS INDIVIDUAIS.xlsx'
        }
        
        self.colunas_mapeamento = {
            'status': ['SITUACAO', 'STATUS', 'SITUAÇÃO', 'ESTADO'],
            'data': ['DATA', 'DATA_CONCLUSAO', 'DATA_INICIO', 'DT_CONCLUSAO'],
            'responsavel': ['RESPONSAVEL', 'COLABORADOR', 'NOME'],
            'prioridade': ['PRIORIDADE', 'URGENCIA', 'IMPORTANCIA'],
            'tipo': ['TIPO', 'CATEGORIA', 'CLASSIFICACAO']
        }
        
        # Status específicos solicitados
        self.status_especificos = [
            'VERIFICADO', 'ANÁLISE', 'PENDENTE', 'PRIORIDADE', 
            'PRIORIDADE TOTAL', 'APROVADO', 'APREENDIDO', 'CANCELADO',
            'QUITADO', 'OUTROS ACORDOS', 'M.ENCAMINHADA'  # Adicionados novos status encontrados
        ]
        
        # Status prioritários (mais importantes)
        self.status_prioritarios = ['APROVADO', 'VERIFICADO']
        
        # Horas de trabalho por dia
        self.horas_trabalho = 8
        
        self.modelos = {}
        warnings.filterwarnings('ignore')
        
    def calcular_metricas_avancadas(self, dados_grupos):
        """Calcula métricas avançadas de produtividade e eficiência"""
        metricas_avancadas = []
        
        for grupo, dados in dados_grupos.items():
            for colaborador, df in dados['colaboradores'].items():
                if df.empty:
                    continue
                
                # Contagem de status
                status_counts = self.contar_status_especificos(df)
                
                # Calcular métricas de produtividade
                total_registros = len(df)
                total_prioritarios = status_counts.get('APROVADO', 0)
                taxa_prioritarios = (total_prioritarios / total_registros * 100) if total_registros > 0 else 0
                
                # Calcular produtividade por hora
                dias_unicos = len(df['DIA'].unique()) if 'DIA' in df.columns else 1
                horas_totais = dias_unicos * self.horas_trabalho
                produtividade_hora = total_registros / horas_totais if horas_totais > 0 else 0
                produtividade_prioritarios = total_prioritarios / horas_totais if horas_totais > 0 else 0
                
                # Calcular tempo médio para aprovação
                tempo_medio = 0
                if 'TEMPO_PROCESSAMENTO' in df.columns:
                    df_concluidos = df[df['STATUS'] == 'APROVADO']
                    if not df_concluidos.empty:
                        tempo_medio = df_concluidos['TEMPO_PROCESSAMENTO'].mean()
                
                # Adicionar métricas
                metricas_avancadas.append({
                    'grupo': grupo,
                    'colaborador': colaborador,
                    'total_registros': total_registros,
                    'aprovados': status_counts.get('APROVADO', 0),
                    'pendentes': status_counts.get('PENDENTE', 0),
                    'taxa_prioritarios': taxa_prioritarios,
                    'produtividade_hora': produtividade_hora,
                    'produtividade_prioritarios': produtividade_prioritarios,
                    'tempo_medio_dias': tempo_medio,
                    'score_eficiencia': (taxa_prioritarios * 0.4 + 
                                        produtividade_hora * 0.3 + 
                                        produtividade_prioritarios * 0.3)
                })
        
        return pd.DataFrame(metricas_avancadas)
    
    def contar_status_especificos(self, df):
        """Conta a ocorrência de cada status específico no DataFrame"""
        if df.empty or 'STATUS' not in df.columns:
            return {}
        
        return {status: len(df[df['STATUS'] == status]) for status in self.status_especificos}
    
    def gerar_ranking_colaboradores(self, df_metricas):
        """Gera ranking de colaboradores baseado em eficiência"""
        if df_metricas.empty:
            return pd.DataFrame()
        
        # Ordenar por score de eficiência
        df_ranking = df_metricas.sort_values('score_eficiencia', ascending=False).reset_index(drop=True)
        df_ranking['ranking'] = df_ranking.index + 1
        
        # Calcular percentil
        df_ranking['percentil'] = df_ranking['score_eficiencia'].rank(pct=True) * 100
        
        return df_ranking
    
    def identificar_melhores_praticas(self, df_metricas, dados_grupos):
        """Identifica as melhores práticas dos colaboradores mais eficientes"""
        if df_metricas.empty:
            return []
        
        # Selecionar top 3 colaboradores
        top_colaboradores = df_metricas.nlargest(3, 'score_eficiencia')
        
        melhores_praticas = []
        
        for _, row in top_colaboradores.iterrows():
            grupo = row['grupo']
            colaborador = row['colaborador']
            
            # Obter dados do colaborador
            df = dados_grupos[grupo]['colaboradores'].get(colaborador)
            if df is None or df.empty:
                continue
            
            # Analisar padrões de sucesso
            # 1. Distribuição por dia da semana
            if 'DIA' in df.columns:
                try:
                    df['DIA'] = pd.to_datetime(df['DIA'])
                    df['DIA_SEMANA'] = df['DIA'].dt.day_name()
                    dias_produtivos = df[df['STATUS'].isin(self.status_prioritarios)]['DIA_SEMANA'].value_counts()
                    if not dias_produtivos.empty:
                        melhor_dia = dias_produtivos.idxmax()
                        melhores_praticas.append(f"O colaborador {colaborador} tem melhor desempenho às {melhor_dia}s")
                except Exception as e:
                    print(f"Aviso: Não foi possível analisar dias da semana para {colaborador}: {str(e)}")
            
            # 2. Verificar se há padrão de horário (se houver coluna de hora)
            hora_cols = [col for col in df.columns if 'HORA' in col]
            if hora_cols:
                try:
                    df_hora = df[df['STATUS'].isin(self.status_prioritarios)]
                    if not df_hora.empty and hora_cols[0] in df_hora.columns:
                        df_hora['HORA'] = pd.to_datetime(df_hora[hora_cols[0]]).dt.hour
                        horas_produtivas = df_hora['HORA'].value_counts()
                        if not horas_produtivas.empty:
                            melhor_hora = horas_produtivas.idxmax()
                            melhores_praticas.append(f"O colaborador {colaborador} é mais produtivo por volta das {melhor_hora}h")
                except Exception as e:
                    print(f"Aviso: Não foi possível analisar horários para {colaborador}: {str(e)}")
            
            # 3. Verificar tipos de contratos com maior taxa de sucesso
            tipo_cols = [col for col in df.columns if any(tipo in col.upper() for tipo in ['TIPO', 'CATEGORIA'])]
            if tipo_cols:
                try:
                    df_tipos = df[df['STATUS'].isin(self.status_prioritarios)]
                    if not df_tipos.empty and tipo_cols[0] in df_tipos.columns:
                        tipos_sucesso = df_tipos[tipo_cols[0]].value_counts()
                        if not tipos_sucesso.empty:
                            melhor_tipo = tipos_sucesso.idxmax()
                            melhores_praticas.append(f"O colaborador {colaborador} tem maior sucesso com contratos do tipo '{melhor_tipo}'")
                except Exception as e:
                    print(f"Aviso: Não foi possível analisar tipos de contrato para {colaborador}: {str(e)}")
        
        return melhores_praticas
    
    def gerar_recomendacoes_estrategicas(self, df_metricas, dados_grupos):
        """Gera recomendações estratégicas para melhorar a produtividade"""
        if df_metricas.empty:
            return []
        
        recomendacoes = []
        
        try:
            # Criar DataFrame consolidado com reset_index para evitar problemas de índice
            todos_dados = []
            for grupo, dados in dados_grupos.items():
                for colaborador, df in dados['colaboradores'].items():
                    if not df.empty:
                        df_temp = df.copy()
                        df_temp['GRUPO'] = grupo
                        df_temp['COLABORADOR'] = colaborador
                        df_temp = df_temp.reset_index(drop=True)
                        todos_dados.append(df_temp)
            
            if todos_dados:
                todos_dados = pd.concat(todos_dados, ignore_index=True)
                
                # Análise de horários produtivos
                if 'HORA' in todos_dados.columns:
                    try:
                        todos_dados['HORA'] = pd.to_datetime(todos_dados['HORA']).dt.hour
                        horas_produtivas = todos_dados[
                            todos_dados['STATUS'].isin(self.status_prioritarios)
                        ]['HORA'].value_counts()
                        
                        if not horas_produtivas.empty:
                            melhores_horas = horas_produtivas.nlargest(2).index.tolist()
                            recomendacoes.append(
                                f"Concentrar esforços de negociação nos horários mais "
                                f"produtivos: {melhores_horas[0]}h e {melhores_horas[1]}h"
                            )
                    except Exception as e:
                        print(f"Aviso: Não foi possível analisar horários produtivos: {str(e)}")
                
                # Análise de dias produtivos
                if 'DIA' in todos_dados.columns:
                    try:
                        todos_dados['DIA'] = pd.to_datetime(todos_dados['DIA'])
                        todos_dados['DIA_SEMANA'] = todos_dados['DIA'].dt.day_name()
                        dias_produtivos = todos_dados[
                            todos_dados['STATUS'].isin(self.status_prioritarios)
                        ]['DIA_SEMANA'].value_counts()
                        
                        if not dias_produtivos.empty:
                            melhor_dia = dias_produtivos.idxmax()
                            recomendacoes.append(
                                f"Priorizar negociações complexas às {melhor_dia}s, "
                                "dia com maior taxa de sucesso"
                            )
                    except Exception as e:
                        print(f"Aviso: Não foi possível analisar dias produtivos: {str(e)}")
            
            # Recomendações baseadas em métricas gerais
            try:
                # Identificar colaboradores com baixa produtividade
                baixa_produtividade = df_metricas[
                    df_metricas['produtividade_hora'] < df_metricas['produtividade_hora'].median()
                ]
                if not baixa_produtividade.empty:
                    recomendacoes.append(
                        "Oferecer treinamento adicional para colaboradores com "
                        "produtividade abaixo da média"
                    )
                
                # Identificar colaboradores com alta taxa de pendências
                if 'pendentes' in df_metricas.columns:
                    alta_pendencia = df_metricas[
                        df_metricas['pendentes'] > df_metricas['pendentes'].median() * 1.5
                    ]
                    if not alta_pendencia.empty:
                        recomendacoes.append(
                            "Redistribuir contratos pendentes de colaboradores sobrecarregados"
                        )
                
                # Identificar grupos com melhor desempenho
                if 'grupo' in df_metricas.columns and 'score_eficiencia' in df_metricas.columns:
                    desempenho_grupos = df_metricas.groupby('grupo')['score_eficiencia'].mean()
                    if not desempenho_grupos.empty:
                        melhor_grupo = desempenho_grupos.idxmax()
                        recomendacoes.append(
                            f"Adotar práticas do grupo {melhor_grupo} que apresenta melhor "
                            "eficiência média"
                        )
            
            except Exception as e:
                print(f"Aviso: Erro ao gerar recomendações baseadas em métricas: {str(e)}")
        
        except Exception as e:
            print(f"Aviso: Erro ao analisar padrões de dados: {str(e)}")
        
        return recomendacoes
    
    def gerar_html_responsivo(self, dados_grupos):
        """
        Gera o dashboard HTML com os dados organizados e validados
        """
        try:
            print("\n=== Iniciando Organização e Validação do Relatório ===")
            
            # DataFrame consolidado para relatório diário
            relatorio_diario = []
            
            for grupo, dados in dados_grupos.items():
                for colaborador, df in dados['colaboradores'].items():
                    if df.empty:
                        print(f"Debug: DataFrame vazio para {colaborador} do grupo {grupo}")
                        continue
                    
                    # Identificar coluna de resolução
                    coluna_resolucao = next(
                        (col for col in df.columns 
                         if 'RESOLUCAO' in col.upper() or 'DATA_CONCLUSAO' in col.upper()),
                        None
                    )
                    
                    if not coluna_resolucao:
                        print(f"Debug: Coluna de resolução não encontrada para {colaborador}")
                        continue
                    
                    try:
                        # Converter e validar datas
                        df['DATA_RESOLUCAO'] = pd.to_datetime(
                            df[coluna_resolucao], 
                            errors='coerce'
                        )
                        
                        # Remover registros com datas inválidas
                        registros_invalidos = df['DATA_RESOLUCAO'].isna().sum()
                        if registros_invalidos > 0:
                            print(f"Debug: {registros_invalidos} registros com data inválida para {colaborador}")
                        
                        df = df.dropna(subset=['DATA_RESOLUCAO'])
                        
                        # Agrupar por data e status
                        daily_stats = df.groupby([
                            df['DATA_RESOLUCAO'].dt.date,
                            'STATUS'
                        ]).size().reset_index(name='quantidade')
                        
                        # Criar estrutura para o dashboard
                        for _, row in daily_stats.iterrows():
                            relatorio_diario.append({
                                'data': row['DATA_RESOLUCAO'],
                                'colaborador': colaborador,
                                'grupo': grupo,
                                'status': row['STATUS'],
                                'quantidade': row['quantidade']
                            })
                        
                    except Exception as e:
                        print(f"Debug: Erro ao processar dados de {colaborador}: {str(e)}")
                        continue
            
            # Converter para DataFrame
            df_relatorio = pd.DataFrame(relatorio_diario)
            
            # Validar estrutura final
            if df_relatorio.empty:
                print("Erro: Nenhum dado válido para o relatório diário")
                return None
            
            # Validar campos necessários para o dashboard
            campos_obrigatorios = ['data', 'colaborador', 'grupo', 'status', 'quantidade']
            campos_faltantes = set(campos_obrigatorios) - set(df_relatorio.columns)
            if campos_faltantes:
                print(f"Erro: Campos obrigatórios faltantes: {campos_faltantes}")
                return None
            
            # Preparar dados para o dashboard
            dados_dashboard = {
                'dados_diarios': df_relatorio.to_dict('records'),
                'resumo_status': df_relatorio.groupby('status')['quantidade'].sum().to_dict(),
                'colaboradores': df_relatorio['colaborador'].unique().tolist(),
                'grupos': df_relatorio['grupo'].unique().tolist(),
                'datas_disponiveis': df_relatorio['data'].dt.date.unique().tolist()
            }
            
            # Validar funcionalidade dos botões e menus
            self.validar_elementos_dashboard(dados_dashboard)
            
            print("\n=== Validação do Relatório Concluída ===")
            print(f"✓ Total de registros: {len(df_relatorio)}")
            print(f"✓ Colaboradores: {len(dados_dashboard['colaboradores'])}")
            print(f"✓ Grupos: {len(dados_dashboard['grupos'])}")
            print(f"✓ Período: {min(dados_dashboard['datas_disponiveis'])} a {max(dados_dashboard['datas_disponiveis'])}")
            
            return dados_dashboard
            
        except Exception as e:
            print(f"\nErro crítico na organização do relatório: {str(e)}")
            traceback.print_exc()
            return None

    def validar_elementos_dashboard(self, dados_dashboard):
        """
        Valida a funcionalidade dos elementos interativos do dashboard
        """
        try:
            print("\n=== Validando Elementos do Dashboard ===")
            
            # 1. Validar filtros
            elementos = {
                'Filtro de Grupos': dados_dashboard['grupos'],
                'Filtro de Colaboradores': dados_dashboard['colaboradores'],
                'Filtro de Datas': dados_dashboard['datas_disponiveis']
            }
            
            for elemento, dados in elementos.items():
                if not dados:
                    print(f"⚠️ {elemento}: Sem dados disponíveis")
                else:
                    print(f"✓ {elemento}: {len(dados)} opções disponíveis")
            
            # 2. Validar dados para gráficos
            if not dados_dashboard['resumo_status']:
                print("⚠️ Dados insuficientes para gráficos de status")
            else:
                print(f"✓ Dados de status disponíveis: {len(dados_dashboard['resumo_status'])} categorias")
            
            # 3. Gerar JavaScript de teste
            js_teste = {
                'filtroGrupo': f"const grupos = {json.dumps(dados_dashboard['grupos'])};",
                'filtroColaborador': f"const colaboradores = {json.dumps(dados_dashboard['colaboradores'])};",
                'filtroData': f"const datas = {json.dumps([str(d) for d in dados_dashboard['datas_disponiveis']])};",
                'dadosGraficos': f"const dadosStatus = {json.dumps(dados_dashboard['resumo_status'])};"
            }
            
            # Salvar JavaScript de teste
            with open('dashboard_test.js', 'w', encoding='utf-8') as f:
                for key, value in js_teste.items():
                    f.write(f"// Teste de {key}\n{value}\n\n")
            
            print("✓ Arquivo de teste JavaScript gerado: dashboard_test.js")
            
        except Exception as e:
            print(f"Erro na validação dos elementos: {str(e)}")

    def gerar_relatorio_txt(self, dados_grupos):
        """Gera um arquivo TXT com todos os relatórios"""
        diretorio = self.encontrar_diretorio()
        arquivo_saida = diretorio / 'relatorio_completo.txt'
        
        with open(arquivo_saida, 'w', encoding='utf-8') as f:
            # Cabeçalho
            f.write("=" * 50 + "\n")
            f.write("RELATÓRIO COMPLETO DE ANÁLISE DE CONTRATOS\n")
            f.write(f"Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
            f.write("=" * 50 + "\n\n")
            
            for grupo, dados in dados_grupos.items():
                f.write(f"GRUPO: {grupo.upper()}\n")
                f.write("=" * 50 + "\n\n")
                
                for colaborador, metricas in dados['metricas'].items():
                    # Relatório Diário
                    f.write(f"{colaborador}\n")
                    f.write("-" * 50 + "\n")
                    f.write("RELATÓRIO DIÁRIO\n")
                    f.write("-" * 30 + "\n")
                    for status in self.status_especificos:
                        count = metricas['metricas_diarias'].get(status, 0)
                        f.write(f"{status:<20} {count:>5}\n")
                    f.write("-" * 30 + "\n")
                    f.write(f"{'TOTAL':<20} {metricas['total_registros']:>5}\n\n")
                    
                    # Relatório Geral
                    f.write("RELATÓRIO GERAL\n")
                    f.write("-" * 30 + "\n")
                    for status in self.status_especificos:
                        count = metricas['status_counts'].get(status, 0)
                        f.write(f"{status:<20} {count:>5}\n")
                    f.write("-" * 30 + "\n")
                    f.write(f"{'TOTAL':<20} {metricas['total_registros']:>5}\n\n")
                    
                    # Métricas de Produtividade
                    f.write("MÉTRICAS DE PRODUTIVIDADE\n")
                    f.write("-" * 30 + "\n")
                    prod_hora = metricas['produtividade_hora']
                    prod_diaria = prod_hora * self.horas_trabalho
                    f.write(f"{'Prod. Diária:':<20} {prod_diaria:>5.1f}\n")
                    f.write(f"{'Prod. Horária:':<20} {prod_hora:>5.2f}\n")
                    
                    status_counts = metricas['status_counts']
                    status_prioritarios = {s: status_counts.get(s, 0) for s in self.status_prioritarios}
                    total_prioritarios = sum(status_prioritarios.values())
                    eficiencia = (total_prioritarios/metricas['total_registros']*100 
                                if metricas['total_registros'] > 0 else 0)
                    
                    f.write(f"{'Eficiência (%):':<20} {eficiencia:>5.1f}\n")
                    f.write("\n" + "=" * 50 + "\n\n")
                
                # Totais do Grupo
                f.write(f"TOTAIS DO GRUPO {grupo.upper()}\n")
                f.write("-" * 50 + "\n")
                totais_grupo = {status: 0 for status in self.status_especificos}
                total_grupo = 0
                
                for _, metricas in dados['metricas'].items():
                    for status in self.status_especificos:
                        totais_grupo[status] += metricas['status_counts'].get(status, 0)
                    total_grupo += metricas['total_registros']
                
                for status, total in totais_grupo.items():
                    f.write(f"{status:<20} {total:>5}\n")
                f.write("-" * 30 + "\n")
                f.write(f"{'TOTAL GERAL':<20} {total_grupo:>5}\n\n")
        
        print(f"\nRelatório TXT gerado em: {arquivo_saida}")
        return arquivo_saida

    def executar_analise_completa(self):
        """Executa o fluxo completo de análise"""
        print("=== Iniciando Análise Inteligente de Desempenho ===\n")
        
        try:
            # 1. Carregar dados
            dados_grupos = self.carregar_dados()
            
            # 2. Validar dados antes de prosseguir
            if not self.validar_dados_antes_geracao(dados_grupos):
                print("\n⚠️ Análise interrompida devido a erros na validação!")
                return False
            
            # 3. Gerar relatório diário
            self.gerar_relatorio_diario(dados_grupos)
            
            # 4. Gerar relatório geral
            self.gerar_relatorio_geral(dados_grupos)
            
            # 5. Gerar relatório de produtividade diária
            self.gerar_relatorio_produtividade_diaria(dados_grupos)
            
            # 6. Gerar relatório TXT
            self.gerar_relatorio_txt(dados_grupos)
            
            # 7. Calcular métricas avançadas
            df_metricas = self.calcular_metricas_avancadas(dados_grupos)
            
            # 8. Gerar ranking de colaboradores
            df_ranking = self.gerar_ranking_colaboradores(df_metricas)
            
            # 9. Identificar melhores práticas
            melhores_praticas = self.identificar_melhores_praticas(df_ranking, dados_grupos)
            
            # 10. Gerar recomendações estratégicas
            recomendacoes = self.gerar_recomendacoes_estrategicas(df_ranking, dados_grupos)
            
            # 11. Gerar relatório HTML responsivo com os dados completos
            self.gerar_html_responsivo(dados_grupos)
            
            print("\n=== Análise Concluída com Sucesso ===")
            return True
            
        except Exception as e:
            print(f"\nErro durante a análise: {str(e)}")
            traceback.print_exc()
            return False

    def encontrar_diretorio(self):
        """Encontra o primeiro diretório válido"""
        for diretorio in self.diretorios:
            if diretorio.exists():
                return diretorio
        raise FileNotFoundError("Nenhum diretório válido encontrado")

    def carregar_dados(self):
        """Carrega e normaliza dados dos arquivos Excel"""
        try:
            diretorio = self.encontrar_diretorio()
            print(f"Diretório de trabalho: {diretorio}")
            
            dados_grupos = {}
            
            for grupo, arquivo in self.arquivos.items():
                caminho = diretorio / arquivo
                print(f"\nTentando carregar arquivo: {caminho}")
                
                if not caminho.exists():
                    print(f"Arquivo não encontrado: {caminho}")
                    continue
                
                print(f"\nCarregando dados do grupo {grupo}...")
                
                try:
                    # Tentar carregar com openpyxl primeiro
                    try:
                        print("Tentando carregar com openpyxl...")
                        wb = load_workbook(filename=str(caminho), read_only=True, data_only=True)
                        abas = [aba for aba in wb.sheetnames if aba not in ["TESTE", "RELATÓRIO GERAL"]]
                        print(f"Abas encontradas: {abas}")
                        
                        dados_colaboradores = {}
                        metricas_colaboradores = {}
                        
                        for aba in abas:
                            try:
                                print(f"\nProcessando aba: {aba}")
                                ws = wb[aba]
                                
                                # Converter worksheet para DataFrame
                                data = []
                                headers = []
                                first_row = True
                                
                                for row in ws.rows:
                                    if first_row:
                                        headers = [str(cell.value).strip() if cell.value else f"Col{i}" for i, cell in enumerate(row)]
                                        first_row = False
                                    else:
                                        data.append([cell.value for cell in row])
                                
                                df = pd.DataFrame(data, columns=headers)
                                
                                if df.empty:
                                    print(f"Aba {aba} está vazia, pulando...")
                                    continue
                                
                                print(f"Colunas encontradas: {df.columns.tolist()}")
                                
                                # Normalizar colunas
                                df.columns = [self.normalizar_valor(col) for col in df.columns]
                                print(f"Colunas após normalização: {df.columns.tolist()}")
                                
                                # Identificar coluna de status
                                col_status = next((col for col in df.columns if any(s in col for s in ['SITUACAO', 'STATUS', 'SITUAÇÃO'])), None)
                                
                                if col_status:
                                    print(f"Coluna de status encontrada: {col_status}")
                                    df = df.rename(columns={col_status: 'STATUS'})
                                    df['STATUS'] = df['STATUS'].fillna('PENDENTE').astype(str).str.upper().str.strip()
                                    
                                    # Mapear status
                                    mapeamento_status = {
                                        'VERIFICADO': 'VERIFICADO',
                                        'ANÁLISE': 'ANÁLISE',
                                        'ANALISE': 'ANÁLISE',
                                        'PENDENTE': 'PENDENTE',
                                        'PRIORIDADE': 'PRIORIDADE',
                                        'PRIORIDADE TOTAL': 'PRIORIDADE TOTAL',
                                        'APROVADO': 'APROVADO',
                                        'APREENDIDO': 'APREENDIDO',
                                        'CANCELADO': 'CANCELADO',
                                        'QUITADO': 'QUITADO',
                                        'OUTROS ACORDOS': 'OUTROS ACORDOS',
                                        'M.ENCAMINHADA': 'M.ENCAMINHADA'
                                    }
                                    df['STATUS'] = df['STATUS'].map(lambda x: mapeamento_status.get(x, x))
                                    
                                    # Adicionar data de processamento
                                    df['DIA'] = pd.to_datetime('today').date()
                                    
                                    dados_colaboradores[aba] = df
                                    
                                    # Calcular métricas básicas
                                    status_counts = df['STATUS'].value_counts().to_dict()
                                    
                                    # Calcular métricas diárias
                                    metricas_diarias = {
                                        'VERIFICADO': 0,
                                        'ANÁLISE': 0,
                                        'PENDENTE': 0,
                                        'PRIORIDADE': 0,
                                        'PRIORIDADE TOTAL': 0,
                                        'APROVADO': 0,
                                        'APREENDIDO': 0,
                                        'CANCELADO': 0,
                                        'QUITADO': 0,
                                        'OUTROS ACORDOS': 0,
                                        'M.ENCAMINHADA': 0
                                    }
                                    metricas_diarias.update(status_counts)
                                    
                                    metricas_colaboradores[aba] = {
                                        'total_registros': len(df),
                                        'status_counts': status_counts,
                                        'metricas_diarias': metricas_diarias,
                                        'produtividade_hora': len(df) / self.horas_trabalho
                                    }
                                    
                                    # Mostrar contagem de status
                                    print(f"✓ {aba}: {len(df)} registros")
                                    print("  Status:")
                                    for status, count in metricas_diarias.items():
                                        print(f"    {status}: {count}")
                                else:
                                    print(f"⚠️ Nenhuma coluna de status encontrada na aba {aba}")
                                  
                            except Exception as e:
                                print(f"✗ Erro ao processar {aba}: {str(e)}")
                                print("Detalhes do erro:")
                                traceback.print_exc()
                                continue
                        
                        wb.close()
                        
                    except Exception as e:
                        print(f"Erro ao processar com openpyxl: {str(e)}")
                        print("Tentando com xlrd...")
                        
                        # Se falhar com openpyxl, tentar com xlrd
                        wb = xlrd.open_workbook(str(caminho))
                        abas = [aba for aba in wb.sheet_names() if aba not in ["TESTE", "RELATÓRIO GERAL"]]
                        
                        dados_colaboradores = {}
                        metricas_colaboradores = {}
                        
                        for aba in abas:
                            try:
                                ws = wb.sheet_by_name(aba)
                                headers = [str(ws.cell_value(0, i)).strip() for i in range(ws.ncols)]
                                data = []
                                
                                for row_idx in range(1, ws.nrows):
                                    row = [ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)]
                                    data.append(row)
                                
                                df = pd.DataFrame(data, columns=headers)
                                
                                # Resto do processamento igual ao anterior...
                                if df.empty:
                                    continue
                                    
                                df.columns = [self.normalizar_valor(col) for col in df.columns]
                                col_status = next((col for col in df.columns if any(s in col for s in ['SITUACAO', 'STATUS', 'SITUAÇÃO'])), None)
                                
                                if col_status:
                                    df = df.rename(columns={col_status: 'STATUS'})
                                    df['STATUS'] = df['STATUS'].fillna('PENDENTE').astype(str).str.upper().str.strip()
                                    
                                    mapeamento_status = {
                                        'VERIFICADO': 'VERIFICADO',
                                        'ANÁLISE': 'ANÁLISE',
                                        'ANALISE': 'ANÁLISE',
                                        'PENDENTE': 'PENDENTE',
                                        'PRIORIDADE': 'PRIORIDADE',
                                        'PRIORIDADE TOTAL': 'PRIORIDADE TOTAL',
                                        'APROVADO': 'APROVADO',
                                        'APREENDIDO': 'APREENDIDO',
                                        'CANCELADO': 'CANCELADO',
                                        'QUITADO': 'QUITADO',
                                        'OUTROS ACORDOS': 'OUTROS ACORDOS',
                                        'M.ENCAMINHADA': 'M.ENCAMINHADA'
                                    }
                                    df['STATUS'] = df['STATUS'].map(lambda x: mapeamento_status.get(x, x))
                                    
                                    df['DIA'] = pd.to_datetime('today').date()
                                    dados_colaboradores[aba] = df
                                    
                                    status_counts = df['STATUS'].value_counts().to_dict()
                                    metricas_diarias = {status: 0 for status in self.status_especificos}
                                    metricas_diarias.update(status_counts)
                                    
                                    metricas_colaboradores[aba] = {
                                        'total_registros': len(df),
                                        'status_counts': status_counts,
                                        'metricas_diarias': metricas_diarias,
                                        'produtividade_hora': len(df) / self.horas_trabalho
                                    }
                                    
                                    print(f"✓ {aba}: {len(df)} registros")
                                    for status, count in metricas_diarias.items():
                                        print(f"    {status}: {count}")
                            
                            except Exception as e:
                                print(f"Erro ao processar {aba} com xlrd: {str(e)}")
                                continue
                    
                    dados_grupos[grupo] = {
                        'colaboradores': dados_colaboradores,
                        'metricas': metricas_colaboradores
                    }
                    
                except Exception as e:
                    print(f"Erro ao processar arquivo {arquivo}: {str(e)}")
                    print("Detalhes do erro:")
                    traceback.print_exc()
                    continue
            
            if not dados_grupos:
                print("\nAtenção: Nenhum dado foi carregado!")
                return {}
            
            print("\nDados carregados com sucesso!")
            return dados_grupos
            
        except Exception as e:
            print(f"\nErro crítico ao carregar dados: {str(e)}")
            print("Detalhes do erro:")
            traceback.print_exc()
            return {}

    def normalizar_valor(self, valor):
        if isinstance(valor, (int, float)):
            return str(valor)
        elif isinstance(valor, str):
            return valor.strip().upper()
        else:
            return str(valor)

    def gerar_relatorio_diario(self, dados_grupos):
        """Gera relatório diário com contagem de status específicos"""
        print("\n=== Relatório Diário ===")
        
        for grupo, dados in dados_grupos.items():
            print(f"\nGrupo: {grupo.upper()}")
            
            for colaborador, metricas in dados['metricas'].items():
                print(f"\n{colaborador}:")
                print("  RELATÓRIO DIÁRIO")
                print("  " + "-" * 30)
                for status in self.status_especificos:
                    count = metricas['metricas_diarias'].get(status, 0)
                    print(f"  {status:<20} {count:>5}")
                print("  " + "-" * 30)
                print(f"  {'TOTAL':<20} {metricas['total_registros']:>5}")
                print()

    def gerar_relatorio_geral(self, dados_grupos):
        """Gera relatório geral com totais por status e colaborador"""
        print("\n=== Relatório Geral ===")
        
        totais_gerais = {status: 0 for status in self.status_especificos}
        total_geral_registros = 0
        
        for grupo, dados in dados_grupos.items():
            print(f"\nGrupo: {grupo.upper()}")
            
            totais_grupo = {status: 0 for status in self.status_especificos}
            total_grupo_registros = 0
            
            # Exibir dados por colaborador
            for colaborador, metricas in dados['metricas'].items():
                print(f"\n{colaborador}:")
                print("  RELATÓRIO GERAL")
                print("  " + "-" * 30)
                
                for status in self.status_especificos:
                    count = metricas['status_counts'].get(status, 0)
                    totais_grupo[status] += count
                    totais_gerais[status] += count
                    print(f"  {status:<20} {count:>5}")
                
                print("  " + "-" * 30)
                total_colaborador = metricas['total_registros']
                total_grupo_registros += total_colaborador
                total_geral_registros += total_colaborador
                print(f"  {'TOTAL':<20} {total_colaborador:>5}")
                print()
            
            # Exibir totais do grupo
            print(f"\nTOTAL DO GRUPO {grupo.upper()}:")
            print("  " + "-" * 30)
            for status, total in totais_grupo.items():
                print(f"  {status:<20} {total:>5}")
            print("  " + "-" * 30)
            print(f"  {'TOTAL GERAL':<20} {total_grupo_registros:>5}")
        
        # Exibir totais gerais
        print("\n=== TOTAIS CONSOLIDADOS ===")
        print("-" * 30)
        for status, total in totais_gerais.items():
            print(f"{status:<20} {total:>5}")
        print("-" * 30)
        print(f"{'TOTAL GERAL':<20} {total_geral_registros:>5}")
        
        print("\nRelatório geral concluído!")

    def gerar_relatorio_produtividade_diaria(self, dados_grupos):
        """Gera relatório de produtividade diária por colaborador"""
        print("\n=== Relatório de Produtividade Diária ===")
        
        for grupo, dados in dados_grupos.items():
            print(f"\nGrupo: {grupo.upper()}")
            
            for colaborador, metricas in dados['metricas'].items():
                # Calcular produtividade por hora
                prod_hora = metricas['produtividade_hora']
                prod_diaria = prod_hora * self.horas_trabalho
                
                # Calcular produtividade por status
                status_counts = metricas['status_counts']
                status_prioritarios = {s: status_counts.get(s, 0) for s in self.status_prioritarios}
                total_prioritarios = sum(status_prioritarios.values())
                
                # Exibir resultados
                print(f"\n{colaborador}:")
                print(f"  Produtividade diária: {prod_diaria:.1f} registros/dia")
                print(f"  Produtividade horária: {prod_hora:.2f} registros/hora")
                print(f"  Status prioritários ({', '.join(self.status_prioritarios)}): {total_prioritarios}")
                print(f"  Eficiência: {(total_prioritarios/metricas['total_registros']*100 if metricas['total_registros'] > 0 else 0):.1f}%")
        
        print("\nRelatório de produtividade concluído!")

    def treinar_modelo_predicao(self, dados_grupos):
        """Treina modelo de machine learning para prever probabilidade de aprovação de contratos"""
        # Preparar dados para treinamento
        X_data = []
        y_data = []
        
        for grupo, dados in dados_grupos.items():
            for colaborador, df in dados['colaboradores'].items():
                if df.empty:
                    continue
                
                # Extrair features
                for _, row in df.iterrows():
                    features = []
                    
                    # Adicionar features numéricas
                    if 'TEMPO_PROCESSAMENTO' in df.columns:
                        features.append(row.get('TEMPO_PROCESSAMENTO', 0))
                    else:
                        features.append(0)
                    
                    # Adicionar hora do dia (se disponível)
                    if 'HORA' in df.columns:
                        features.append(pd.to_datetime(row['HORA']).hour)
                    else:
                        features.append(12)  # valor padrão
                    
                    # Adicionar dia da semana (se disponível)
                    if 'DIA' in df.columns:
                        features.append(pd.to_datetime(row['DIA']).weekday())
                    else:
                        features.append(0)
                    
                    # Target: 1 se aprovado/quitado, 0 caso contrário
                    status = row.get('STATUS', '')
                    y_data.append(1 if status in self.status_prioritarios else 0)
                    X_data.append(features)
        
        if not X_data or not y_data:
            return None
        
        # Converter para arrays numpy
        X = np.array(X_data)
        y = np.array(y_data)
        
        # Normalizar features
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        
        # Dividir dados em treino e teste
        X_train, X_test, y_train, y_test = train_test_split(
            X_scaled, y, test_size=0.2, random_state=42
        )
        
        # Treinar modelo
        model = RandomForestClassifier(
            n_estimators=100,
            max_depth=None,
            min_samples_split=2,
            random_state=42
        )
        
        model.fit(X_train, y_train)
        
        # Avaliar modelo
        y_pred = model.predict(X_test)
        accuracy = (y_pred == y_test).mean()
        
        print("\n=== Modelo de Predição de Aprovação ===")
        print(f"Acurácia do modelo: {accuracy:.2%}")
        
        # Salvar modelo e scaler
        self.modelos['predicao'] = {
            'model': model,
            'scaler': scaler,
            'accuracy': accuracy,
            'features': ['tempo_processamento', 'hora_dia', 'dia_semana']
        }
        
        return model

    def prever_aprovacao(self, tempo_processamento, hora_dia=None, dia_semana=None):
        """Prevê a probabilidade de aprovação de um contrato"""
        if 'predicao' not in self.modelos:
            return None
            
        modelo = self.modelos['predicao']
        
        # Preparar features
        features = [
            tempo_processamento,
            hora_dia if hora_dia is not None else 12,
            dia_semana if dia_semana is not None else 0
        ]
        
        # Normalizar features
        X = modelo['scaler'].transform([features])
        
        # Fazer predição
        prob = modelo['model'].predict_proba(X)[0]
        
        return {
            'probabilidade_aprovacao': prob[1],
            'confianca_modelo': modelo['accuracy']
        }

    def validar_dados_antes_geracao(self, dados_grupos):
        """
        Valida os dados antes de gerar relatórios, tratando status faltantes de forma inteligente.
        """
        try:
            validacoes = {
                'grupos': True,
                'colaboradores': True,
                'metricas': True,
                'status': True,
                'relatorios': True
            }
            
            # Estrutura para armazenar status faltantes por colaborador
            status_faltantes_por_colaborador = {}
            
            # 1. Primeira passagem: Identificar status realmente utilizados
            status_em_uso = set()
            for grupo, dados in dados_grupos.items():
                for colaborador, metricas in dados['metricas'].items():
                    if 'status_counts' in metricas:
                        status_em_uso.update(metricas['status_counts'].keys())
            
            # 2. Validação principal com tratamento inteligente de status
            for grupo, dados in dados_grupos.items():
                for colaborador, metricas in dados['metricas'].items():
                    if 'status_counts' not in metricas:
                        continue
                    
                    # Identificar status faltantes
                    status_colaborador = set(metricas['status_counts'].keys())
                    status_faltantes = set(self.status_especificos) - status_colaborador
                    
                    # Filtrar apenas status relevantes
                    status_faltantes_relevantes = {
                        status for status in status_faltantes
                        if status in status_em_uso  # Status é usado por outros colaboradores
                    }
                    
                    if status_faltantes_relevantes:
                        status_faltantes_por_colaborador[colaborador] = status_faltantes_relevantes
                    
                    # Adicionar status faltantes com valor zero
                    for status in status_faltantes:
                        metricas['status_counts'][status] = 0
                        metricas['metricas_diarias'][status] = 0
            
            # 3. Análise de impacto e decisão
            if status_faltantes_por_colaborador:
                print("\n=== Análise de Status Faltantes ===")
                status_criticos = {'APROVADO', 'PENDENTE', 'VERIFICADO'}  # Status essenciais
                tem_status_critico = False
                
                for colaborador, status_faltantes in status_faltantes_por_colaborador.items():
                    status_criticos_faltantes = status_faltantes & status_criticos
                    if status_criticos_faltantes:
                        tem_status_critico = True
                        print(f"⚠️ {colaborador}: Faltam status críticos: {status_criticos_faltantes}")
                    else:
                        print(f"ℹ️ {colaborador}: Status não críticos ausentes: {status_faltantes}")
                
                # Decidir se interrompe baseado apenas em status críticos
                validacoes['status'] = not tem_status_critico
            
            # 4. Relatório de validação
            print("\n=== Resultado da Validação ===")
            for tipo, valido in validacoes.items():
                status = "✓" if valido else "✗"
                print(f"{status} {tipo.capitalize()}")
            
            if not all(validacoes.values()):
                print("\n⚠️ Atenção:")
                if not validacoes['status']:
                    print("• Foram encontrados status críticos faltantes")
                    print("• Os dados podem estar incompletos ou incorretos")
                return False
            
            print("\n✓ Validação concluída com sucesso!")
            print("• Status não críticos faltantes foram preenchidos com zero")
            print("• Os relatórios podem ser gerados com segurança")
            return True

        except Exception as e:
            print(f"\nErro durante a validação: {str(e)}")
            traceback.print_exc()
            return False

    def exportar_para_sqlite(self, dados_grupos=None):
        """Exporta os dados para um banco de dados SQLite"""
        try:
            print("\n=== Exportando dados para SQLite ===")
            
            # Criar instância do gerenciador de banco de dados
            db_manager = RelatorioDatabase()
            
            # Verificar se já existe relatório gerado
            relatorio_path = Path('F:/relatoriotest/relatorio_completo.txt')
            if relatorio_path.exists():
                print(f"Importando dados do relatório: {relatorio_path}")
                if db_manager.importar_relatorio_txt(relatorio_path):
                    print("✓ Dados exportados com sucesso!")
                    return True
                else:
                    print("✗ Falha ao importar relatório.")
                    return False
            else:
                print("✗ Arquivo de relatório não encontrado.")
                return False
        
        except Exception as e:
            print(f"Erro ao exportar para SQLite: {str(e)}")
            traceback.print_exc()
            return False

    def gerar_dashboard_sqlite(self):
        """Gera um dashboard HTML a partir do banco de dados SQLite"""
        try:
            print("\n=== Gerando Dashboard a partir do SQLite ===")
            
            # Conectar ao banco de dados
            db_path = "F:/relatoriotest/relatorio_dashboard.db"
            conn = sqlite3.connect(db_path)
            
            # Consultas SQL otimizadas
            query_status = """
            SELECT 
                SUM(verificado) as verificado,
                SUM(analise) as analise,
                SUM(pendente) as pendente,
                SUM(prioridade) as prioridade,
                SUM(prioridade_total) as prioridade_total,
                SUM(aprovado) as aprovado,
                SUM(apreendido) as apreendido,
                SUM(cancelado) as cancelado
            FROM relatorio_geral
            """
            
            # Executar consultas
            df_status = pd.read_sql_query(query_status, conn)
            
            # Verificar se há dados válidos e substituir NaN por 0
            df_status = df_status.fillna(0)
            
            # Gerar gráfico de pizza
            plt.figure(figsize=(10, 6))
            status_labels = ['Verificado', 'Análise', 'Pendente', 'Prioridade', 
                             'Prioridade Total', 'Aprovado', 'Apreendido', 'Cancelado']
            status_values = [
                df_status['verificado'].iloc[0], 
                df_status['analise'].iloc[0],
                df_status['pendente'].iloc[0], 
                df_status['prioridade'].iloc[0],
                df_status['prioridade_total'].iloc[0], 
                df_status['aprovado'].iloc[0],
                df_status['apreendido'].iloc[0], 
                df_status['cancelado'].iloc[0]
            ]
            
            # Verificar se há pelo menos um valor não zero
            if sum(status_values) > 0:
                plt.pie(status_values, labels=status_labels, autopct='%1.1f%%')
                plt.title('Distribuição de Status')
            else:
                # Se não houver dados, criar um gráfico vazio
                plt.text(0.5, 0.5, 'Sem dados disponíveis', 
                        horizontalalignment='center', verticalalignment='center')
                plt.axis('off')
            
            # Resto do código permanece igual...
            
            # Gerar HTML
            template_str = '''
            <!DOCTYPE html>
            <html lang="pt-br">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>Dashboard de Produtividade</title>
                <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
                <style>
                    .dashboard-card { margin-bottom: 20px; }
                    .chart-container { margin: 20px 0; }
                </style>
            </head>
            <body>
                <div class="container mt-4">
                    <h1 class="text-center mb-4">Dashboard de Produtividade</h1>
                    
                    <!-- Status -->
                    <div class="row dashboard-card">
                        <div class="col-12">
                            <div class="card">
                                <div class="card-header">
                                    <h4>Status dos Contratos</h4>
                                </div>
                                <div class="card-body">
                                    <div class="row">
                                        <div class="col-md-6">
                                            <img src="data:image/png;base64,{{ grafico_pizza }}" class="img-fluid">
                                        </div>
                                        <div class="col-md-6">
                                            <table class="table">
                                                <thead>
                                                    <tr>
                                                        <th>Status</th>
                                                        <th>Quantidade</th>
                                                    </tr>
                                                </thead>
                                                <tbody>
                                                    {% for status, valor in status_data.items() %}
                                                    <tr>
                                                        <td>{{ status }}</td>
                                                        <td>{{ valor }}</td>
                                                    </tr>
                                                    {% endfor %}
                                                </tbody>
                                            </table>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                    
                    <!-- Eficiência -->
                    <div class="row dashboard-card">
                        <div class="col-12">
                            <div class="card">
                                <div class="card-header">
                                    <h4>Eficiência por Colaborador</h4>
                                </div>
                                <div class="card-body">
                                    <img src="data:image/png;base64,{{ grafico_barras }}" class="img-fluid">
                                </div>
                            </div>
                        </div>
                    </div>
                    
                    <!-- Métricas -->
                    <div class="row dashboard-card">
                        <div class="col-12">
                            <div class="card">
                                <div class="card-header">
                                    <h4>Métricas Detalhadas</h4>
                                </div>
                                <div class="card-body">
                                    <div class="table-responsive">
                                        <table class="table table-striped">
                                            <thead>
                                                <tr>
                                                    <th>Colaborador</th>
                                                    <th>Grupo</th>
                                                    <th>Total</th>
                                                    <th>Prod. Diária</th>
                                                    <th>Prod. Horária</th>
                                                    <th>Eficiência</th>
                                                </tr>
                                            </thead>
                                            <tbody>
                                                {% for _, row in df_metricas.iterrows() %}
                                                <tr>
                                                    <td>{{ row.colaborador }}</td>
                                                    <td>{{ row.grupo }}</td>
                                                    <td>{{ row.total }}</td>
                                                    <td>{{ "%.2f"|format(row.prod_diaria) }}</td>
                                                    <td>{{ "%.2f"|format(row.prod_horaria) }}</td>
                                                    <td>{{ "%.2f"|format(row.eficiencia) }}%</td>
                                                </tr>
                                                {% endfor %}
                                            </tbody>
                                        </table>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
                
                <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>
            </body>
            </html>
            '''
            
            # Renderizar template
            template = jinja2.Template(template_str)
            html_output = template.render(
                status_data=df_status.iloc[0].to_dict(),
                grafico_pizza=base64.b64encode(buffer.read()).decode('utf-8'),
                grafico_barras=base64.b64encode(buffer.read()).decode('utf-8'),
                df_metricas=df_metricas
            )
            
            # Salvar HTML
            output_path = Path('F:/relatoriotest/dashboard.html')
            output_path.write_text(html_output, encoding='utf-8')
            
            print(f"✓ Dashboard gerado com sucesso: {output_path}")
            return True
            
        except Exception as e:
            print(f"Erro ao gerar dashboard: {str(e)}")
            traceback.print_exc()
            return False
        finally:
            if 'conn' in locals():
                conn.close()

class RelatorioDatabase:
    def __init__(self, db_path="F:/relatoriotest/relatorio_dashboard.db"):
        self.db_path = db_path
        self.conn = None
        self.criar_tabelas()
        self.horas_trabalho = 8  # Horas de trabalho por dia

    def conectar(self):
        """Estabelece conexão com o banco de dados"""
        self.conn = sqlite3.connect(self.db_path)
        return self.conn
    
    def fechar(self):
        """Fecha a conexão com o banco de dados"""
        if self.conn:
            self.conn.close()
    
    def criar_tabelas(self):
        """Cria as tabelas necessárias no banco de dados"""
        conn = self.conectar()
        cursor = conn.cursor()
        
        # Tabela de grupos
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS grupos (
            id INTEGER PRIMARY KEY,
            nome TEXT UNIQUE NOT NULL
        )
        ''')
        
        # Tabela de colaboradores
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS colaboradores (
            id INTEGER PRIMARY KEY,
            nome TEXT NOT NULL,
            grupo_id INTEGER NOT NULL,
            FOREIGN KEY (grupo_id) REFERENCES grupos (id),
            UNIQUE (nome, grupo_id)
        )
        ''')
        
        # Tabela de relatório geral
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS relatorio_geral (
            id INTEGER PRIMARY KEY,
            colaborador_id INTEGER NOT NULL,
            data_relatorio DATE NOT NULL,
            verificado INTEGER DEFAULT 0,
            analise INTEGER DEFAULT 0,
            pendente INTEGER DEFAULT 0,
            prioridade INTEGER DEFAULT 0,
            prioridade_total INTEGER DEFAULT 0,
            aprovado INTEGER DEFAULT 0,
            apreendido INTEGER DEFAULT 0,
            cancelado INTEGER DEFAULT 0,
            total INTEGER DEFAULT 0,
            FOREIGN KEY (colaborador_id) REFERENCES colaboradores (id),
            UNIQUE (colaborador_id, data_relatorio)
        )
        ''')
        
        # Tabela de métricas de produtividade
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS metricas_produtividade (
            id INTEGER PRIMARY KEY,
            colaborador_id INTEGER NOT NULL,
            data_relatorio DATE NOT NULL,
            prod_diaria REAL DEFAULT 0,
            prod_horaria REAL DEFAULT 0,
            eficiencia REAL DEFAULT 0,
            FOREIGN KEY (colaborador_id) REFERENCES colaboradores (id),
            UNIQUE (colaborador_id, data_relatorio)
        )
        ''')
        
        conn.commit()
        self.fechar()

    def importar_relatorio_txt(self, caminho_relatorio):
        """Importa dados do relatório txt para o banco de dados"""
        try:
            with open(caminho_relatorio, 'r', encoding='utf-8') as file:
                conteudo = file.read()
            
            # Obter data de geração do relatório
            data_match = re.search(r'Data de Geração: (\d{2}/\d{2}/\d{4})', conteudo)
            data_relatorio = datetime.now().date()
            if data_match:
                try:
                    data_relatorio = datetime.strptime(data_match.group(1), "%d/%m/%Y").date()
                except:
                    pass
            
            conn = self.conectar()
            cursor = conn.cursor()
            
            # Extrair grupos e colaboradores
            grupos = {}
            grupo_atual = None
            colaborador_atual = None
            dados_relatorio = []
            dados_metricas = []
            
            for linha in conteudo.split('\n'):
                linha = linha.strip()
                
                # Identificar grupo
                if linha.startswith('GRUPO: '):
                    grupo_atual = linha.replace('GRUPO: ', '').strip()
                    # Inserir grupo se não existir
                    cursor.execute('INSERT OR IGNORE INTO grupos (nome) VALUES (?)', (grupo_atual,))
                    cursor.execute('SELECT id FROM grupos WHERE nome = ?', (grupo_atual,))
                    grupos[grupo_atual] = cursor.fetchone()[0]
                
                # Identificar colaborador (linhas que não começam com espaço e não são cabeçalhos)
                elif not linha.startswith(' ') and linha and not linha.startswith('=') and not linha.startswith('-') and not linha.startswith('GRUPO:') and not linha.startswith('RELATÓRIO'):
                    colaborador_atual = linha
                    # Inserir colaborador se não existir
                    if grupo_atual and colaborador_atual:
                        cursor.execute('''
                            INSERT OR IGNORE INTO colaboradores (nome, grupo_id)
                            VALUES (?, ?)
                        ''', (colaborador_atual, grupos[grupo_atual]))
                        cursor.execute('SELECT id FROM colaboradores WHERE nome = ? AND grupo_id = ?',
                                     (colaborador_atual, grupos[grupo_atual]))
                        colaborador_id = cursor.fetchone()[0]
                        
                        # Inicializar dados do relatório para este colaborador
                        dados_colaborador = {
                            'colaborador_id': colaborador_id,
                            'data_relatorio': data_relatorio,
                            'verificado': 0,
                            'analise': 0,
                            'pendente': 0,
                            'prioridade': 0,
                            'prioridade_total': 0,
                            'aprovado': 0,
                            'apreendido': 0,
                            'cancelado': 0
                        }
                        dados_relatorio.append(dados_colaborador)
                
                # Processar linhas de status
                elif colaborador_atual and ':' in linha:
                    status, valor = linha.split(':')
                    status = status.strip().upper()
                    try:
                        valor = int(valor.strip())
                        if status in ['VERIFICADO', 'ANÁLISE', 'PENDENTE', 'PRIORIDADE', 'PRIORIDADE TOTAL',
                                    'APROVADO', 'APREENDIDO', 'CANCELADO']:
                            dados_colaborador[status.lower()] = valor
                    except ValueError:
                        continue
            
            # Inserir dados no relatório geral
            for dados in dados_relatorio:
                total = sum(dados[k] for k in ['verificado', 'analise', 'pendente', 'prioridade',
                                             'prioridade_total', 'aprovado', 'apreendido', 'cancelado'])
                
                cursor.execute('''
                    INSERT OR REPLACE INTO relatorio_geral (
                        colaborador_id, data_relatorio,
                        verificado, analise, pendente,
                        prioridade, prioridade_total,
                        aprovado, apreendido, cancelado,
                        total
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    dados['colaborador_id'],
                    dados['data_relatorio'],
                    dados['verificado'],
                    dados['analise'],
                    dados['pendente'],
                    dados['prioridade'],
                    dados['prioridade_total'],
                    dados['aprovado'],
                    dados['apreendido'],
                    dados['cancelado'],
                    total
                ))
                
                # Calcular métricas de produtividade
                prod_horaria = total / self.horas_trabalho if self.horas_trabalho > 0 else 0
                prod_diaria = prod_horaria * self.horas_trabalho
                eficiencia = (dados['aprovado'] / total * 100) if total > 0 else 0
                
                # Inserir métricas de produtividade
                cursor.execute('''
                    INSERT OR REPLACE INTO metricas_produtividade (
                        colaborador_id, data_relatorio,
                        prod_diaria, prod_horaria, eficiencia
                    ) VALUES (?, ?, ?, ?, ?)
                ''', (
                    dados['colaborador_id'],
                    dados['data_relatorio'],
                    prod_diaria,
                    prod_horaria,
                    eficiencia
                ))
            
            conn.commit()
            print(f"✓ Dados importados com sucesso para o banco de dados: {self.db_path}")
            return True
        
        except Exception as e:
            print(f"Erro ao importar relatório: {str(e)}")
            traceback.print_exc()
            return False
        finally:
            self.fechar()

if __name__ == "__main__":
    try:
        print("\n=== Iniciando Análise de Dados ===")
        analisador = AnalisadorInteligente()
        print("\nExecutando análise completa...")
        analisador.executar_analise_completa()
        print("\n=== Análise Concluída com Sucesso ===")
        analisador.exportar_para_sqlite()
        analisador.gerar_dashboard_sqlite()
    except Exception as e:
        print(f"\nErro crítico: {str(e)}")
        print("Detalhes do erro:")
        traceback.print_exc() 